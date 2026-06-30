"""Transforma los archivos que sube el cliente (JSON o Excel/zip) al dataset del
Tablero Financiero — el **Paso 2** (curar/automatizar) hecho dentro de MaestroAI.

- Si se sube un ``.json`` con el dataset curado, se usa tal cual (mezclado sobre el
  demo para garantizar que existan todas las llaves del contrato).
- Si se suben los Excel (Resumen por proyecto, Concentrado BC, Timesheet + catálogo de
  horas, Evaluación de clientes) se derivan los bloques reales: proyectos (P&L por año,
  mezcla de costo, EBITDA real vs BC), costo por hora, utilización, comparativo de
  costos (costo_bc; cmi/timesheet quedan pendientes hasta Nómina), evaluación y Gob/IP.

Las partes que no vienen en los Excel de proyecto (estados financieros por entidad,
balance) se conservan del dataset base y se marca ``partial_entities`` hasta que se
suba un JSON completo o se conecte la fuente (Paso 1).
"""
from __future__ import annotations

import datetime
import io
import json
import re
import zipfile
from collections import defaultdict
from copy import deepcopy

from . import dataset as _dataset

_MES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]


def _num(x) -> float:
    return float(x) if isinstance(x, (int, float)) and not isinstance(x, bool) else 0.0


def _expand(files: list[tuple[str, bytes]]) -> list[tuple[str, bytes]]:
    """Descomprime cualquier .zip y devuelve la lista plana de archivos."""
    out: list[tuple[str, bytes]] = []
    for name, raw in files:
        if name.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(raw)) as z:
                    for n in z.namelist():
                        if n.endswith("/") or "__MACOSX" in n or n.startswith("."):
                            continue
                        out.append((n.split("/")[-1], z.read(n)))
            except zipfile.BadZipFile:
                continue
        else:
            out.append((name, raw))
    return out


def _wb(raw: bytes):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)


def _year_of(name: str) -> str:
    digits = "".join(c for c in name if c.isdigit())
    for i in range(len(digits) - 3):
        chunk = digits[i:i + 4]
        if chunk.startswith("20"):
            return chunk
    return ""


# ----- Resumen por proyecto -----
_RC = dict(id=1, cliente=3, nombre=4, venta=30, costos=31, hwsw=32, nomina=33, grep=34,
           corp=35, amex=36, viat=37, caja=38, uber=39, otros=40, actint=41, noreg=42,
           margen=44, ebitda=47, ebitda_bc=51, sector=53)


def _resumen(files):
    years, projects_by_year = {}, {}
    for name, raw in files:
        if "resumen" not in name.lower() or not name.lower().endswith(".xlsx"):
            continue
        yr = _year_of(name)
        if not yr:
            continue
        try:
            wb = _wb(raw)
        except Exception:
            continue
        ws = wb[wb.sheetnames[0]]
        agg = dict(venta=0, costos=0, margen=0, ebitda=0, ebitda_bc=0, gob=0, ip=0, proyectos=0,
                   nomina=0, hwsw=0, corp=0, repr_viat=0, otros=0)
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if len(r) < 54 or not isinstance(r[_RC['id']], (int, float)) or isinstance(r[_RC['id']], bool):
                continue
            venta = _num(r[_RC['venta']])
            if venta <= 0:
                continue
            gob = str(r[_RC['sector']] or "").strip().upper().startswith("GOB")
            margen, ebitda, ebitda_bc = _num(r[_RC['margen']]), _num(r[_RC['ebitda']]), _num(r[_RC['ebitda_bc']])
            agg['venta'] += venta; agg['costos'] += _num(r[_RC['costos']]); agg['margen'] += margen
            agg['ebitda'] += ebitda; agg['ebitda_bc'] += ebitda_bc; agg['proyectos'] += 1
            agg['nomina'] += _num(r[_RC['nomina']]); agg['hwsw'] += _num(r[_RC['hwsw']])
            agg['corp'] += _num(r[_RC['corp']])
            agg['repr_viat'] += sum(_num(r[_RC[k]]) for k in ('grep', 'amex', 'viat', 'caja', 'uber'))
            agg['otros'] += sum(_num(r[_RC[k]]) for k in ('otros', 'actint', 'noreg'))
            agg['gob' if gob else 'ip'] += venta
            rows.append(dict(cliente=str(r[_RC['cliente']] or "").strip(),
                             nombre=str(r[_RC['nombre']] or "").strip()[:60],
                             tipo="Gobierno" if gob else "Privado", venta=round(venta),
                             costos=round(_num(r[_RC['costos']])), margen=round(margen),
                             pct_margen=round(margen / venta, 3) if venta else 0,
                             ebitda=round(ebitda), ebitda_bc=round(ebitda_bc),
                             desviacion=round(ebitda - ebitda_bc)))
        wb.close()
        if agg['venta'] > 1_000_000:
            years[yr] = agg
            projects_by_year[yr] = sorted(rows, key=lambda x: -x['venta'])
    return years, projects_by_year


def _agg_clients(rows):
    d = {}
    for p in rows:
        c = d.setdefault(p['cliente'], dict(name=p['cliente'], sector=p['tipo'], revenue=0, _m=0))
        c['revenue'] += p['venta']; c['_m'] += p['margen']
    out = []
    for c in d.values():
        c['margin'] = round(c['_m'] / c['revenue'], 3) if c['revenue'] else 0
        c['status'] = "red" if c['margin'] < 0 else ("amber" if c['margin'] < 0.1 else "green")
        c.pop('_m'); out.append(c)
    return sorted(out, key=lambda x: -x['revenue'])


# ----- Concentrado BC: costo_bc mensual + costo-hora por rol -----
def _concentrado(files):
    by_year_month, roles = {}, {}
    for name, raw in files:
        low = name.lower()
        if "concentrado" not in low or not low.endswith(".xlsx"):
            continue
        yr = _year_of(name)
        try:
            wb = _wb(raw)
        except Exception:
            continue
        ws = wb[[s for s in wb.sheetnames if "Concentrado_BC" == s] and "Concentrado_BC" or wb.sheetnames[-1]]
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        iRol = next((i for i, h in enumerate(hdr) if h and "Recursos humanos" in str(h)), None)
        iCxh = next((i for i, h in enumerate(hdr) if h and "Costo x hora" in str(h)), None)
        months = [0.0] * 12
        for r in ws.iter_rows(min_row=2, values_only=True):
            if len(r) < 43:
                continue
            for j, c in enumerate(range(31, 43)):
                months[j] += _num(r[c])
            if iRol is not None and iCxh is not None and iRol < len(r) and iCxh < len(r):
                rol = str(r[iRol]).strip() if r[iRol] else None
                cxh = _num(r[iCxh])
                if rol and cxh > 0:
                    roles.setdefault(rol, []).append(cxh)
        wb.close()
        if sum(months) > 1000 and yr:
            by_year_month[yr] = [round(m) for m in months]
    costo_hora = sorted([dict(rol=k, costo_hora=round(sum(v) / len(v)), registros=len(v))
                         for k, v in roles.items()], key=lambda x: -x['costo_hora'])[:12]
    return by_year_month, costo_hora


# ----- Timesheet: detectado por columnas (Empleado/Proyecto/Total de Horas), no por
# nombre de archivo — así no depende de que el cliente lo llame "timesheet*.xlsx" -----
def _parse_fecha(v) -> tuple[str, int | None]:
    """Acepta texto 'DD/MM/YYYY' o un date/datetime de openpyxl → (año, mes)."""
    if isinstance(v, (datetime.date, datetime.datetime)):
        return str(v.year), v.month
    s = str(v or "").strip().replace("-", "/")
    parts = s.split("/")
    if len(parts) == 3:
        d, m, y = parts
        if len(y) == 4 and y.isdigit() and m.isdigit():
            return y, int(m)
    return "", None


def _parse_timesheet(files):
    """Suma horas por (año), (año,mes) y (año,mes,proyecto) + empleados/proyecto por
    año, cruzando cualquier hoja .xlsx que tenga columnas Empleado/Proyecto/Total de
    Horas (insensible al nombre del archivo, para soportar exports de cualquier sistema
    de timesheet)."""
    hours_by_year: dict[str, float] = defaultdict(float)
    hours_by_year_month: dict[tuple[str, int], float] = defaultdict(float)
    emps_by_year: dict[str, set] = defaultdict(set)
    proj_by_year: dict[tuple[str, str], float] = defaultdict(float)
    found = False
    for name, raw in files:
        if not name.lower().endswith(".xlsx"):
            continue
        try:
            wb = _wb(raw)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            try:
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue

            def idx(label):
                return next((i for i, h in enumerate(hdr) if h and label.lower() in str(h).lower()), None)
            iE, iP, iH, iF = idx("Empleado"), idx("Proyecto"), idx("Total de Horas"), idx("Fecha")
            if iE is None or iP is None or iH is None:
                continue
            found = True
            for r in ws.iter_rows(min_row=2, values_only=True):
                h = _num(r[iH]) if iH < len(r) else 0
                if h <= 0:
                    continue
                emp = str(r[iE]).strip() if iE < len(r) and r[iE] else ""
                proj = str(r[iP]).strip().split("-")[0].strip()[:24] if iP < len(r) and r[iP] else "Otros"
                yr, mm = _parse_fecha(r[iF]) if iF is not None and iF < len(r) else ("", None)
                if not yr:
                    continue
                hours_by_year[yr] += h
                if emp:
                    emps_by_year[yr].add(emp)
                proj_by_year[(yr, proj)] += h
                if mm:
                    hours_by_year_month[(yr, mm)] += h
        wb.close()
    return dict(found=found, hours_by_year=hours_by_year, hours_by_year_month=hours_by_year_month,
                emps_by_year=emps_by_year, proj_by_year=proj_by_year)


def _capacidad_horas(files) -> dict[str, float]:
    cap: dict[str, float] = {}
    for name, raw in files:
        if "horas laborales" in name.lower():
            try:
                wb = _wb(raw)
            except Exception:
                continue
            ws = wb[wb.sheetnames[0]]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0] and r[2]:
                    cap[str(int(r[0]))] = cap.get(str(int(r[0])), 0) + _num(r[2])
            wb.close()
    return cap


def _utilizacion(ts: dict, cap: dict) -> dict | None:
    """Bloque ``utilization`` (año con más horas) a partir de lo que sumó ``_parse_timesheet``."""
    if not ts["found"] or not ts["hours_by_year"]:
        return None
    yr = max(ts["hours_by_year"], key=lambda y: ts["hours_by_year"][y])
    tot = ts["hours_by_year"][yr]
    emps = ts["emps_by_year"].get(yr, set())
    capy = cap.get(yr) or (sum(cap.values()) / len(cap) if cap else 2000)
    capacidad = capy * len(emps)
    proj = {p: h for (y, p), h in ts["proj_by_year"].items() if y == yr}
    return dict(year=yr, horas_reales=round(tot), empleados=len(emps),
                horas_capacidad=round(capacidad), capacidad_emp=round(capy),
                utilizacion=round(tot / capacidad, 3) if capacidad else 0,
                by_project=[dict(nombre=n, horas=round(hh))
                            for n, hh in sorted(proj.items(), key=lambda x: -x[1])[:8]])


# ----- Nómina (lista de raya, p. ej. export CONTPAQi): costo_cmi -----
_PERC_COL = "*TOTAL* *PERCEPCIONES*"
_OBL_COL = "*TOTAL* *OBLIGACIONES*"
_PERIODO_RE = re.compile(r"(\d{2}/\d{2}/\d{4})\s*al\s*(\d{2}/\d{2}/\d{4})")


def _nomina(files) -> dict:
    """Detecta hojas de Nómina por sus columnas *TOTAL* *PERCEPCIONES* / *TOTAL*
    *OBLIGACIONES* (percepciones + obligaciones patronales = costo real a la empresa).
    Un mismo archivo puede traer varias hojas (una por año). Devuelve
    {año: {costo_cmi, empleados, meses: [(año, mes), ...]}}."""
    out: dict[str, dict] = {}
    for name, raw in files:
        if not name.lower().endswith(".xlsx"):
            continue
        try:
            wb = _wb(raw)
        except Exception:
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            head = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))
            hdr_row = next((r for r in head if r and _PERC_COL in r and _OBL_COL in r), None)
            if hdr_row is None:
                continue
            iPerc, iObl = hdr_row.index(_PERC_COL), hdr_row.index(_OBL_COL)
            m = _PERIODO_RE.search(" ".join(str(c) for r in head for c in r if c))
            meses: list[tuple[str, int]] = []
            if m:
                d0 = datetime.datetime.strptime(m.group(1), "%d/%m/%Y")
                d1 = datetime.datetime.strptime(m.group(2), "%d/%m/%Y")
                yy, mm = d0.year, d0.month
                while (yy, mm) <= (d1.year, d1.month):
                    meses.append((str(yy), mm))
                    mm += 1
                    if mm > 12:
                        mm, yy = 1, yy + 1
            yr = (meses[0][0] if meses else "") or _year_of(sn) or _year_of(name)
            if not yr:
                continue
            total, n = 0.0, 0
            for r in ws.iter_rows(min_row=1, values_only=True):
                cod = r[0] if r else None
                if not isinstance(cod, str) or not cod.strip().isdigit():
                    continue
                total += _num(r[iPerc]) + _num(r[iObl])
                n += 1
            if total > 0:
                prev = out.get(yr)
                if not prev or total > prev["costo_cmi"]:
                    out[yr] = dict(costo_cmi=round(total), empleados=n,
                                   meses=meses or [(yr, mo) for mo in range(1, 13)])
        wb.close()
    return out


# ----- Evaluación de clientes -----
def _scoring(files):
    for name, raw in files:
        if "evalua" not in name.lower() or not name.lower().endswith(".xlsx"):
            continue
        try:
            wb = _wb(raw)
        except Exception:
            continue
        sn = next((s for s in wb.sheetnames if "Evaluaci" in s and "vf" in s), None) or \
            next((s for s in wb.sheetnames if "Evaluaci" in s), None)
        if not sn:
            wb.close(); continue
        ws = wb[sn]
        clients = []
        for r in ws.iter_rows(min_row=3, values_only=True):
            if len(r) < 28 or not r[2] or not isinstance(r[26], (int, float)):
                continue
            clients.append(dict(name=str(r[2]).strip(), sector=str(r[4] or "").strip(),
                                score=round(float(r[26]), 1), tier=str(r[27] or "").strip(),
                                facturacion=str(r[11] or "").strip(), rentabilidad=str(r[20] or "").strip()))
        wb.close()
        if clients:
            return dict(criteria=[["Monto de facturación", 0.40], ["Recurrencia", 0.15],
                                  ["Antigüedad", 0.15], ["Rentabilidad", 0.15],
                                  ["Días de pago", 0.10], ["Tamaño de empresa", 0.05]],
                        clients=sorted(clients, key=lambda c: -c['score']))
    return None


def build_dataset_from_files(files: list[tuple[str, bytes]]) -> dict:
    """Construye el dataset desde lo que suba el cliente. ``files`` = [(nombre, bytes)]."""
    flat = _expand(files)
    names = ", ".join(n for n, _ in files)[:200]

    # 1) Si hay un JSON, es el contrato completo: úsalo (sobre el demo para llenar huecos).
    for n, raw in flat:
        if n.lower().endswith(".json"):
            base = deepcopy(_dataset.load())
            base.pop("_origin", None); base.pop("_is_demo", None)
            try:
                base.update(json.loads(raw.decode("utf-8")))
            except Exception as exc:
                raise ValueError(f"JSON inválido: {exc}")
            base["_source_files"] = names
            return base

    # 2) Excel: derivar los bloques reales sobre el dataset base.
    years, projects_by_year = _resumen(flat)
    bc_months, costo_hora = _concentrado(flat)
    nomina_by_year = _nomina(flat)
    ts = _parse_timesheet(flat)
    sc = _scoring(flat)
    if not (years or bc_months or nomina_by_year or ts["found"] or sc):
        raise ValueError("No se reconoció ningún archivo válido (Resumen por proyecto, "
                         "Concentrado BC, Nómina, Timesheet, Evaluación) ni un .json del dataset.")
    out = deepcopy(_dataset.load())
    out.pop("_origin", None); out.pop("_is_demo", None)

    latest = None
    if years:
        # Año "actual" = el de mayor venta (cierre principal), no el último del calendario
        # (que suele ser parcial/forecast).
        latest = max(years, key=lambda y: years[y]['venta'])
        trend = {y: dict(venta=round(a['venta']), gob=round(a['gob']), ip=round(a['ip']),
                         ebitda=round(a['ebitda']), ebitda_bc=round(a['ebitda_bc']),
                         margen=round(a['margen']), proyectos=a['proyectos'],
                         desviacion=round(a['ebitda'] - a['ebitda_bc']))
                 for y, a in years.items()}
        a = years[latest]
        out["projects"] = dict(
            source=f"Cargado desde Excel ({names})",
            totals=dict(venta=round(a['venta']), costos=round(a['costos']), margen=round(a['margen']),
                        ebitda=round(a['ebitda']), ebitda_bc=round(a['ebitda_bc']),
                        pct_margen=round(a['margen'] / a['venta'], 3) if a['venta'] else 0,
                        pct_ebitda=round(a['ebitda'] / a['venta'], 3) if a['venta'] else 0,
                        desviacion=round(a['ebitda'] - a['ebitda_bc']), proyectos=a['proyectos']),
            trend=trend,
            cost_mix=dict(nomina=round(a['nomina']), hw_sw=round(a['hwsw']), costo_corp=round(a['corp']),
                          repr_viaticos=round(a['repr_viat']), otros=round(a['otros'])),
            clients=_agg_clients(projects_by_year[latest])[:20],
            detail=projects_by_year[latest][:30])
        # Gob/IP real desde la tendencia
        out["gob_ip"] = {y: dict(gob=trend[y]['gob'], ip=trend[y]['ip']) for y in trend}

    if costo_hora:
        out["cost_per_hour"] = dict(year=latest or (max(nomina_by_year) if nomina_by_year else
                                                     out.get("cost_per_hour", {}).get("year", "")),
                                    by_role=costo_hora)

    # ----- cost_comparison: costo_bc (Concentrado) + costo_cmi (Nómina, prorrateado por
    # mes dentro del periodo reportado) + costo_timesheet (horas reales × costo-hora
    # promedio del periodo = costo_cmi del periodo ÷ horas del periodo) -----
    by_month_map: dict[tuple[str, int], dict] = {}
    for yr in sorted(bc_months):
        for j, mm in enumerate(_MES):
            by_month_map[(yr, j + 1)] = dict(anio=yr, mes=mm, costo_bc=bc_months[yr][j],
                                             costo_cmi=None, costo_timesheet=None)

    def _rec(myr: str, mm: int) -> dict:
        key = (myr, mm)
        rec = by_month_map.get(key)
        if rec is None:
            rec = dict(anio=myr, mes=_MES[mm - 1], costo_bc=None, costo_cmi=None, costo_timesheet=None)
            by_month_map[key] = rec
        return rec

    cmi_years, ts_years = [], []
    for yr, info in nomina_by_year.items():
        meses = info["meses"]
        monto_mes = info["costo_cmi"] / (len(meses) or 12)
        cmi_years.append(yr)
        for myr, mm in meses:
            _rec(myr, mm)["costo_cmi"] = round(monto_mes)

    if ts["found"]:
        for yr, info in nomina_by_year.items():
            meses = info["meses"]
            monto_mes = info["costo_cmi"] / (len(meses) or 12)
            # Solo los meses donde el Timesheet sí registró horas (puede arrancar a
            # mitad del periodo de Nómina) — comparar costo de esos meses contra esas
            # horas, no el costo del año completo contra horas de unos pocos meses.
            meses_con_horas = [(myr, mm) for myr, mm in meses if ts["hours_by_year_month"].get((myr, mm), 0) > 0]
            horas_periodo = sum(ts["hours_by_year_month"][k] for k in meses_con_horas)
            if not meses_con_horas or horas_periodo <= 0:
                continue
            costo_hora_prom = (monto_mes * len(meses_con_horas)) / horas_periodo
            ts_years.append(yr)
            for myr, mm in meses:
                h = ts["hours_by_year_month"].get((myr, mm))
                if h:
                    _rec(myr, mm)["costo_timesheet"] = round(h * costo_hora_prom)

    if by_month_map:
        available = [k for k, on in (("costo_bc", bool(bc_months)), ("costo_cmi", bool(cmi_years)),
                                     ("costo_timesheet", bool(ts_years))) if on]
        pending = [k for k in ("costo_bc", "costo_cmi", "costo_timesheet") if k not in available]
        notes = []
        if bc_months:
            notes.append("costo_bc derivado de Concentrado BC.")
        if cmi_years:
            notes.append(f"costo_cmi de Nómina ({', '.join(sorted(cmi_years))}; percepciones + "
                         "obligaciones patronales), prorrateado por mes dentro del periodo reportado.")
        if ts_years:
            notes.append("costo_timesheet = horas reales del Timesheet × costo-hora promedio del "
                         "periodo (costo_cmi del periodo ÷ horas del periodo).")
        if pending:
            notes.append(f"Pendiente: {', '.join(pending)}.")
        out["cost_comparison"] = dict(note=" ".join(notes), available=available, pending=pending,
                                      by_month=sorted(by_month_map.values(),
                                                      key=lambda r: (r["anio"], _MES.index(r["mes"]))))

    util = _utilizacion(ts, _capacidad_horas(flat))
    if util:
        out["utilization"] = util
    if sc:
        out["client_scoring"] = sc

    if years:
        out["company"] = {**out.get("company", {}), "period": f"Cargado desde Excel · cierre {latest}"}
    elif nomina_by_year:
        cierre = max(nomina_by_year)
        out["company"] = {**out.get("company", {}),
                          "period": f"Cargado desde Excel (Nómina/Timesheet) · cierre {cierre}"}
    out["partial_entities"] = True  # estados financieros por entidad pendientes (requieren EF/JSON)
    out["_source_files"] = names
    return out

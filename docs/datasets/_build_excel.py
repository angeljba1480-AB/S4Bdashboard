# -*- coding: utf-8 -*-
"""Construye palabras-regionales.xlsx combinando:
- ES ancho (una columna por pais) desde _gen_wide_es.DATA + workflow_words.json (si existe)
- USA slang y UK slang
Hojas: 'ES (ancho)', 'ES (largo)', 'USA slang', 'UK slang', 'LEEME'
Uso: python3 _build_excel.py [workflow_words.json]
"""
import csv, os, sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from _gen_wide_es import DATA as ES_DATA, PAISES, ALL

OUT = os.path.dirname(os.path.abspath(__file__))

# Cargar palabras extra del workflow si se pasa un JSON
extra = []
if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
    raw = json.load(open(sys.argv[1], encoding="utf-8"))
    words = raw.get("words", raw) if isinstance(raw, dict) else raw
    for w in words:
        variantes = [(v.get("sig",""), [c.strip() for c in v.get("paises","").split(",") if c.strip()])
                     for v in w.get("variantes",[])]
        extra.append((w["palabra"], w.get("emoji",""), w.get("categoria","Varios"),
                      w.get("general",""), variantes))

# Combinar + dedupe por palabra
combined = {}
for w in list(ES_DATA) + extra:
    combined[w[0].strip().lower()] = w
DATA = list(combined.values())

def meanings(w):
    palabra, emoji, cat, general, variantes = w
    asign = {}
    for sig, codes in variantes:
        for c in codes:
            if c in PAISES: asign[c] = sig
    return {c: asign.get(c, general) for c in ALL}

wb = Workbook()
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(color="FFFFFF", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")

def style_header(ws, ncols):
    for j in range(1, ncols+1):
        c = ws.cell(1, j); c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

# --- Hoja ES ancho ---
ws = wb.active; ws.title = "ES (ancho)"
header = ["#", "Palabra", "Emoji", "Categoria"] + [f"{PAISES[c][0]} {PAISES[c][1]}" for c in ALL]
ws.append(header)
for i, w in enumerate(sorted(DATA, key=lambda x: x[0].lower()), 1):
    m = meanings(w)
    row = [i, w[0], w[1], w[2]] + [m[c] for c in ALL]
    ws.append(row)
    if "⚠️" in "".join(str(x) for x in row):
        for j in range(1, len(row)+1): ws.cell(i+1, j).fill = WARN_FILL
style_header(ws, len(header))
ws.column_dimensions['B'].width = 18
for c in ALL:
    ws.column_dimensions[get_column_letter(5+ALL.index(c))].width = 26

# --- Hoja ES largo (palabra x pais) ---
ws2 = wb.create_sheet("ES (largo)")
ws2.append(["Palabra","Emoji","Categoria","Codigo","Pais","Significado"])
for w in sorted(DATA, key=lambda x: x[0].lower()):
    m = meanings(w)
    for c in ALL:
        ws2.append([w[0], w[1], w[2], c, PAISES[c][1], m[c]])
style_header(ws2, 6)
for col,wd in zip("ABCDEF",[18,8,26,8,18,50]): ws2.column_dimensions[col].width = wd

# --- Hojas USA / UK desde CSV ---
def add_csv_sheet(name, csv_name):
    path = os.path.join(OUT, csv_name)
    if not os.path.exists(path): return
    ws = wb.create_sheet(name)
    with open(path, encoding="utf-8") as f:
        for row in csv.reader(f): ws.append(row)
    style_header(ws, ws.max_column)
    for col,wd in zip("ABCDEFGH",[22,10,28,8,48,28,10,24]):
        ws.column_dimensions[col].width = wd
add_csv_sheet("USA slang", "us-slang-regional.csv")
add_csv_sheet("UK slang", "uk-slang-regional.csv")

# --- Hoja LEEME ---
ws3 = wb.create_sheet("LEEME")
info = [
 ["Dataset","Palabras/Modismos del espanol e ingles que cambian por pais/region, con emoji teen"],
 ["",""],
 ["Hoja","Contenido"],
 ["ES (ancho)", f"{len(DATA)} palabras x {len(ALL)} paises hispanohablantes (una columna por pais)"],
 ["ES (largo)", "Misma data en formato largo (una fila por palabra+pais) para tablas dinamicas"],
 ["USA slang", "Modismos de ingles de EE.UU. por region (NYC, South, Midwest, West, AAVE, Gen-Z...)"],
 ["UK slang", "Modismos de ingles de UK por region (London/MLE, Scotland, Wales, Liverpool...)"],
 ["",""],
 ["Leyenda","⚠️ = sentido vulgar/sexual en algun pais (filas resaltadas en naranja)"],
 ["Nota","Material educativo; los significados regionales evolucionan y varian por edad/ciudad/contexto"],
]
for r in info: ws3.append(r)
ws3.cell(1,1).font = Font(bold=True, size=14)
ws3.column_dimensions['A'].width = 16; ws3.column_dimensions['B'].width = 90
wb._sheets.insert(0, wb._sheets.pop())  # LEEME al frente

out = os.path.join(OUT, "palabras-regionales.xlsx")
wb.save(out)
print(f"Excel generado: {out}")
print(f"Hojas: {wb.sheetnames}")
print(f"ES total: {len(DATA)} palabras")
